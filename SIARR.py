import streamlit as st
import pandas as pd
import numpy as np
import time
import mysql.connector
import os
import traceback
import io  
import random 
import hashlib
import re

# Librerías de IA, Formato Excel y Visualización Avanzada
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import tensorflow as tf 
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout, Input
from tensorflow.keras.callbacks import Callback
from openpyxl.styles import PatternFill
import plotly.express as px
import plotly.graph_objects as go

# --- FUNCIONES DE SEGURIDAD (MD5 y Moodle) ---
def generar_md5(texto):
    return hashlib.md5(texto.encode('utf-8')).hexdigest()

def validar_password_moodle(password):
    """Valida: min 8 caracteres, 1 mayúscula, 1 minúscula, 1 número, 1 carácter especial"""
    if len(password) < 8: return False
    if not re.search(r"\d", password): return False
    if not re.search(r"[a-z]", password): return False
    if not re.search(r"[A-Z]", password): return False
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password): return False
    return True

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="SIARR", page_icon="🎓", layout="wide")

# --- CSS PERSONALIZADO (Optimizado para Computadora y Celular) ---
st.markdown("""
<style>
/* Ocultar menú y marca de agua de Streamlit */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* 💻 Diseño para Computadora (Pantallas grandes) */
.block-container {
    padding-top: 2rem !important;
    padding-bottom: 2rem !important;
    padding-left: 3rem !important;
    padding-right: 3rem !important;
    max-width: 100% !important;
}

/* 📱 Diseño Específico para Celulares */
@media (max-width: 768px) {
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 1rem !important;
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
    }
    h1 { font-size: 1.8rem !important; }
    h2 { font-size: 1.5rem !important; }
    h3 { font-size: 1.2rem !important; }
    .stDataFrame {
        overflow-x: auto;
    }
}

/* Botones con estilo moderno */
.stButton>button {
    border-radius: 8px !important;
    font-weight: bold !important;
}
</style>
""", unsafe_allow_html=True)

# --- CONTROL DE SESIONES ---
if 'usuario_actual' not in st.session_state:
    st.session_state['usuario_actual'] = None
if 'rol_actual' not in st.session_state:
    st.session_state['rol_actual'] = None
if 'nombre' not in st.session_state:
    st.session_state['nombre'] = ""
if 'alumno_seleccionado_evaluar' not in st.session_state:
    st.session_state['alumno_seleccionado_evaluar'] = ""
if 'tab_actual' not in st.session_state:
    st.session_state['tab_actual'] = None
if 'analisis_completado' not in st.session_state:
    st.session_state['analisis_completado'] = False
if 'df_resultados' not in st.session_state:
    st.session_state['df_resultados'] = None
if 'excel_data' not in st.session_state:
    st.session_state['excel_data'] = None
if 'df_crudo_entrenamiento' not in st.session_state:
    st.session_state['df_crudo_entrenamiento'] = None 

# Variables para el Dashboard Drill-Down
if 'dash_nivel' not in st.session_state:
    st.session_state['dash_nivel'] = 1
if 'dash_estado_acad' not in st.session_state:
    st.session_state['dash_estado_acad'] = None
if 'dash_semestre' not in st.session_state:
    st.session_state['dash_semestre'] = None
if 'form_alumno_version' not in st.session_state:
    st.session_state['form_alumno_version'] = 0
if 'form_docente_version' not in st.session_state:
    st.session_state['form_docente_version'] = 0
if 'form_alta_usuario_version' not in st.session_state:
    st.session_state['form_alta_usuario_version'] = 0
if 'form_edicion_usuario_version' not in st.session_state:
    st.session_state['form_edicion_usuario_version'] = 0

class StreamlitLogger(Callback):
    def __init__(self, placeholder):
        self.placeholder = placeholder
    def on_epoch_end(self, epoch, logs=None):
        logs = logs or {}
        msg = f"🧠 Entrenando Red Neuronal... Época {epoch+1}/20 | Precisión: {logs.get('accuracy', 0):.2%}"
        self.placeholder.info(msg)

# --- FUNCIÓN DE CONEXIÓN A BASE DE DATOS ---
def get_db_connection():
    if "DB_HOST" in st.secrets:
        return mysql.connector.connect(
            host=st.secrets["DB_HOST"],
            port=int(st.secrets["DB_PORT"]),
            user=st.secrets["DB_USER"],
            password=st.secrets["DB_PASSWORD"],
            database=st.secrets["DB_NAME"]
        )
    else:
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="base_datos_its"
        )

def init_db():
    try:
        if "DB_HOST" not in st.secrets:
            conn = mysql.connector.connect(host="localhost", user="root", password="")
            with conn.cursor() as c:
                c.execute("CREATE DATABASE IF NOT EXISTS base_datos_its")
            conn.close()
            
        with get_db_connection() as conn:
            with conn.cursor() as c:
                c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                            matricula VARCHAR(50) PRIMARY KEY, 
                            password VARCHAR(50), 
                            rol VARCHAR(50), 
                            nombre VARCHAR(100), 
                            correo VARCHAR(100),
                            docente_id VARCHAR(50) DEFAULT NULL)''')
                
                try:
                    c.execute("ALTER TABLE usuarios ADD COLUMN docente_id VARCHAR(50) DEFAULT NULL")
                    conn.commit()
                except:
                    pass

                c.execute("SELECT COUNT(*) FROM usuarios")
                if c.fetchone()[0] == 0:
                    pwd_default = generar_md5('Temporal123*')
                    usuarios_prueba = [
                        ('admin', pwd_default, 'administrative', 'Coordinación General', 'admin@itsperote.edu.mx', None),
                        ('profe_juan', pwd_default, 'docente', 'Ing. Juan Pérez', 'juan@itsperote.edu.mx', None),
                        ('21050002', pwd_default, 'alumno', 'Miguel Angel Sanchez', 'angel@gmail.com', 'profe_juan'),
                        ('21050003', pwd_default, 'alumno', 'Luz Rueda Tereso', 'rueda@gmail.com', 'profe_juan'),
                        ('21050009', pwd_default, 'alumno', 'Miguel Angel Sanchez', 'angel@gmail.com', 'profe_juan')
                    ]
                    c.executemany('INSERT INTO usuarios (matricula, password, rol, nombre, correo, docente_id) VALUES (%s, %s, %s, %s, %s, %s)', usuarios_prueba)
                    conn.commit()

                c.execute('''CREATE TABLE IF NOT EXISTS respuestas_alumnos (
                            matricula VARCHAR(50) PRIMARY KEY,
                            sexo VARCHAR(20), semestre INT, sistema VARCHAR(50), 
                            horas_estudio INT, dias_estudio INT, 
                            motivacion INT, confianza INT, dificultad INT, 
                            apoyo INT, estres INT, 
                            computadora VARCHAR(10), internet VARCHAR(10), calidad_internet INT,
                            FOREIGN KEY (matricula) REFERENCES usuarios(matricula) ON DELETE CASCADE)''')

                c.execute('''CREATE TABLE IF NOT EXISTS evaluaciones_docentes (
                            matricula VARCHAR(50) PRIMARY KEY,
                            promedio FLOAT, reprobadas INT, calif_ultima INT, 
                            dias_asistencia INT, asistencia_clases INT, cumplimiento INT, 
                            participacion INT, practicas INT, uso_plataformas INT,
                            FOREIGN KEY (matricula) REFERENCES usuarios(matricula) ON DELETE CASCADE)''')
                conn.commit()
    except mysql.connector.Error as err:
        st.error(f"Error de inicialización de Base de Datos: {err}.")

if not st.session_state.get('db_inicializada'):
    init_db()
    st.session_state['db_inicializada'] = True

# --- INTERFAZ DE LOGIN ---
if st.session_state['usuario_actual'] is None:
    st.markdown("<br><br><br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.5, 1])
    
    with col2:
        with st.container(border=True):
            st.markdown("<h1 style='text-align: center;'>🎓 Acceso al Sistema</h1>", unsafe_allow_html=True)
            st.write("---")
            
            with st.form("formulario_acceso"):
                usuario = st.text_input("Usuario / Matrícula")
                password = st.text_input("Contraseña", type="password")
                boton_ingresar = st.form_submit_button("Iniciar Sesión", type="primary", use_container_width=True)

                if boton_ingresar:
                    if not usuario or not password:
                        st.error("Introduce tu usuario y contraseña.")
                    else:
                        try:
                            password_cifrado = generar_md5(password)
                            
                            with get_db_connection() as conn:
                                with conn.cursor() as c:
                                    c.execute("SELECT * FROM usuarios WHERE matricula=%s AND password=%s", (usuario, password_cifrado))
                                    resultado = c.fetchone()
                                    
                                    if resultado:
                                        st.session_state['usuario_actual'] = resultado[0]
                                        rol_bd = str(resultado[2]).lower()
                                        if str(resultado[0]).lower() == "admin":
                                            rol_bd = "administrative"
                                        st.session_state['rol_actual'] = rol_bd
                                        st.session_state['nombre'] = resultado[3]
                                        
                                        if 'admin' in rol_bd:
                                            st.session_state['tab_actual'] = "👥 Gestión de Usuarios (CRUD)"
                                        else:
                                            st.session_state['tab_actual'] = "📝 Carga la información del Alumno"
                                            
                                        st.success("¡Acceso concedido!")
                                        st.rerun()
                                    else:
                                        st.error("Usuario o contraseña incorrectos.")
                        except mysql.connector.Error as err:
                            st.error(f"Error en la consulta: {err}")

else:
    # --- MENÚ SUPERIOR DE CIERRE DE SESIÓN ---
    col_usr, col_btn = st.columns([8, 2])
    with col_usr:
        st.markdown(f"**Usuario conectado:** {st.session_state['nombre']} ({st.session_state['rol_actual'].upper()})")
    with col_btn:
        if st.button("❌ Cerrar Sesión", type="secondary", use_container_width=True):
            st.session_state['usuario_actual'] = None
            st.session_state['rol_actual'] = None
            st.session_state['nombre'] = ""
            st.session_state['tab_actual'] = None
            st.session_state['dash_nivel'] = 1
            st.session_state['dash_estado_acad'] = None
            st.session_state['dash_semestre'] = None
            st.rerun()

    with st.expander("👤 Datos personales", expanded=False):
        try:
            with get_db_connection() as conn:
                with conn.cursor() as c:
                    c.execute("SELECT nombre, correo, password FROM usuarios WHERE matricula=%s", (st.session_state['usuario_actual'],))
                    perfil_actual = c.fetchone()

            if perfil_actual:
                col_perfil, col_password = st.columns([1, 1.15])

                with col_perfil:
                    with st.container(border=True):
                        st.markdown("### 👤 Datos personales")
                        st.caption("Actualiza tu nombre y correo institucional.")
                        with st.form("form_datos_personales"):
                            nuevo_nombre = st.text_input("Nombre completo", value=perfil_actual[0] or "")
                            nuevo_correo = st.text_input("Correo electrónico", value=perfil_actual[1] or "")

                            if st.form_submit_button("💾 Guardar datos", type="primary", use_container_width=True):
                                if not nuevo_nombre.strip():
                                    st.error("El nombre no puede estar vacío.")
                                else:
                                    with get_db_connection() as conn:
                                        with conn.cursor() as c:
                                            c.execute("UPDATE usuarios SET nombre=%s, correo=%s WHERE matricula=%s",
                                                      (nuevo_nombre.strip(), nuevo_correo.strip(), st.session_state['usuario_actual']))
                                            conn.commit()
                                    st.session_state['nombre'] = nuevo_nombre.strip()
                                    st.success("Datos personales actualizados correctamente.")
                                    st.rerun()

                with col_password:
                    with st.container(border=True):
                        st.markdown("### 🔐 Cambiar Contraseña de Acceso")
                        st.caption("Modifica tu clave periódicamente para mantener protegidos los archivos y folios de tus proyectos de investigación.")
                        with st.form("form_cambiar_password"):
                            password_actual = st.text_input("Contraseña Actual", type="password", placeholder="Ingresa tu contraseña actual")
                            nueva_password = st.text_input("Nueva Contraseña", type="password", placeholder="Mínimo 8 caracteres")
                            confirmar_password = st.text_input("Confirmar Nueva Contraseña", type="password", placeholder="Repite la nueva contraseña")

                            if st.form_submit_button("🔑 ACTUALIZAR CONTRASEÑA", type="primary", use_container_width=True):
                                if not password_actual or not nueva_password or not confirmar_password:
                                    st.error("Completa los tres campos de contraseña.")
                                elif generar_md5(password_actual) != perfil_actual[2]:
                                    st.error("La contraseña actual no es correcta.")
                                elif nueva_password != confirmar_password:
                                    st.error("La nueva contraseña y la confirmación no coinciden.")
                                elif not validar_password_moodle(nueva_password):
                                    st.error("La nueva contraseña debe tener al menos 8 caracteres, 1 mayúscula, 1 minúscula, 1 número y 1 carácter especial.")
                                else:
                                    with get_db_connection() as conn:
                                        with conn.cursor() as c:
                                            c.execute("UPDATE usuarios SET password=%s WHERE matricula=%s",
                                                      (generar_md5(nueva_password), st.session_state['usuario_actual']))
                                            conn.commit()
                                    st.success("Contraseña actualizada correctamente.")
                                    st.rerun()
            else:
                st.warning("No se encontraron datos del usuario actual.")
        except mysql.connector.Error as err:
            st.error(f"Error al cargar o guardar datos personales: {err}")

    def colorear_filas(row):
        if row['Resultado IA'] == '⚠️ RIESGO':
            return ['background-color: #ffcccc; color: #900000'] * len(row)
        elif row['Resultado IA'] == '✅ ESTABLE':
            return ['background-color: #e6ffe6; color: #006600'] * len(row)
        return [''] * len(row)

    # --- MÓDULO DE IA ---
    def mostrar_modulo_ia():
        st.markdown("### 🧠 Inteligencia Artificial con Aprendizaje Profundo")
        st.write("Detección de riesgo de reprobación en asignaturas de programación mediante análisis de desempeño predictivo.")
        
        if 'admin' in st.session_state['rol_actual']:
            st.markdown("#### 📂 Configuración del Dataset de Entrenamiento")
            col_da1, col_da2 = st.columns([2, 1])
            with col_da1:
                archivo_cargado = st.file_uploader("Agregar un nuevo dataset para el análisis (.csv, .xlsx)", type=["csv", "xlsx"], key="file_uploader_ia")
                if archivo_cargado is not None:
                    try:
                        ext = ".xlsx" if archivo_cargado.name.endswith('.xlsx') else ".csv"
                        nombre_archivo_compartido = f"dataset_compartido_admin{ext}"
                        
                        if os.path.exists("dataset_compartido_admin.xlsx"): os.remove("dataset_compartido_admin.xlsx")
                        if os.path.exists("dataset_compartido_admin.csv"): os.remove("dataset_compartido_admin.csv")
                        
                        with open(nombre_archivo_compartido, "wb") as f:
                            f.write(archivo_cargado.getbuffer())
                            
                        st.success(f"✅ Dataset guardado globalmente en el servidor: {archivo_cargado.name}")
                    except Exception as e:
                        st.error(f"Error al escribir el archivo compartido: {e}")
            with col_da2:
                st.write("")
                st.write("")
                if st.button("🗑️ Borrar Dataset de Análisis", type="secondary", use_container_width=True):
                    if os.path.exists("dataset_compartido_admin.xlsx"): os.remove("dataset_compartido_admin.xlsx")
                    if os.path.exists("dataset_compartido_admin.csv"): os.remove("dataset_compartido_admin.csv")
                    st.session_state['analisis_completado'] = False
                    st.session_state['df_resultados'] = None
                    st.session_state['df_crudo_entrenamiento'] = None
                    st.warning("Dataset personalizado eliminado del servidor. Se usará el archivo local por defecto.")
                    st.rerun()

        if os.path.exists("dataset_compartido_admin.xlsx") or os.path.exists("dataset_compartido_admin.csv"):
            st.info("ℹ️ Actualmente utilizando el dataset compartido cargado por el Administrador.")
        else:
            st.info("ℹ️ Utilizando el dataset histórico predeterminado del sistema institucional.")
            
        if st.button("🚀 Iniciar Análisis de Red Neuronal", type="primary", use_container_width=True):
            consola_placeholder = st.empty()
            consola_placeholder.info("Buscando y cargando archivo de base de datos histórico...")
            
            try:
                os.environ['PYTHONHASHSEED'] = '42'
                np.random.seed(42)
                random.seed(42)
                tf.random.set_seed(42)
                
                if os.path.exists("dataset_compartido_admin.xlsx"):
                    df = pd.read_excel("dataset_compartido_admin.xlsx")
                elif os.path.exists("dataset_compartido_admin.csv"):
                    df = pd.read_csv("dataset_compartido_admin.csv")
                else:
                    archivo_csv = 'dataset_final_sin_duplicados.xlsx - Sheet1.csv'
                    archivo_excel = 'dataset_final_sin_duplicados.xlsx'
                    if os.path.exists(archivo_csv):
                        df = pd.read_csv(archivo_csv)
                    elif os.path.exists(archivo_excel):
                        df = pd.read_excel(archivo_excel)
                    else:
                        st.error("❌ Error: No se encontró ningún archivo de base de datos histórico en el servidor.")
                        return

                # Guardamos copia original para extraer info rica en el Dashboard
                st.session_state['df_crudo_entrenamiento'] = df.copy()

                cols_ignorar = ['Matrícula', 'Matricula', 'matricula', 'Nombre', 'Nombre_completo', 'ID', 'Id']
                df = df.drop(columns=[c for c in cols_ignorar if c in df.columns], errors='ignore')
                
                if 'Resultado' in df.columns:
                    df['Resultado'] = df['Resultado'].map({'Aprobado': 0, 'Reprobado': 1})
                df = df.dropna(subset=['Resultado'])
                
                if 'Sexo' in df.columns:
                    df['Sexo_Num'] = df['Sexo'].map({'Hombre': 0, 'Mujer': 1}).fillna(0)
                if 'Sistema_Escolar' in df.columns:
                    df['Sistema_Escolar_Num'] = df['Sistema_Escolar'].map({'Escolarizado': 0, 'Semiescolarizado': 1}).fillna(0)
                    
                df_train, df_test = train_test_split(df, test_size=0.2, random_state=42)
                y_train = df_train['Resultado']
                X_train_bruto = df_train.drop(columns=['Resultado'], errors='ignore')
                X_train = X_train_bruto.select_dtypes(include=[np.number]).fillna(0)
                
                nombres_variables = X_train.columns
                correlaciones = X_train.corrwith(y_train).fillna(0).values
                
                scaler = StandardScaler()
                X_train_scaled = scaler.fit_transform(X_train)
                
                model = Sequential([
                    Input(shape=(X_train_scaled.shape[1],)),
                    Dense(64, activation='relu'),
                    Dropout(0.3),
                    Dense(32, activation='relu'),
                    Dense(1, activation='sigmoid')
                ])
                
                model.compile(optimizer='adam', loss='binary_crossentropy', metrics=['accuracy'])
                model.fit(X_train_scaled, y_train, epochs=20, shuffle=False, callbacks=[StreamlitLogger(consola_placeholder)], verbose=0)
                consola_placeholder.success("✅ Red neuronal entrenada con éxito.")
                
                query_completos = """
                SELECT 
                    u.matricula, u.nombre, 
                    (SELECT nombre FROM usuarios WHERE matricula = u.docente_id) as maestro_responsable,
                    ra.semestre, ra.sexo, ra.sistema, 
                    ed.promedio, ed.reprobadas, ed.calif_ultima, ed.dias_asistencia, 
                    ed.asistencia_clases, ed.cumplimiento, ed.participacion, ed.practicas, ed.uso_plataformas,
                    ra.horas_estudio, ra.dias_estudio, ra.apoyo, ra.motivacion, ra.confianza, ra.dificultad, ra.estres,
                    ra.computadora, ra.internet, ra.calidad_internet
                FROM usuarios u
                INNER JOIN respuestas_alumnos ra ON u.matricula = ra.matricula
                INNER JOIN evaluaciones_docentes ed ON u.matricula = ed.matricula
                WHERE u.rol = 'alumno'
                """
                
                parametros_query = []
                if st.session_state['rol_actual'] == 'docente':
                    query_completos += " AND u.docente_id = %s"
                    parametros_query.append(st.session_state['usuario_actual'])
                    
                with get_db_connection() as conn:
                    df_db = pd.read_sql(query_completos, conn, params=parametros_query if parametros_query else None)
                    
                if df_db.empty:
                    st.warning("⚠️ No hay alumnos con expedientes completos (deben responder el cuestionario y ser evaluados).")
                    st.session_state['analisis_completado'] = False
                else:
                    df_db_prep = pd.DataFrame()
                    df_db_prep['Matrícula'] = df_db['matricula']
                    df_db_prep['Nombre_completo'] = df_db['nombre']
                    df_db_prep['Docente_Tutor'] = df_db['maestro_responsable'].fillna("Sin Asignar")
                    df_db_prep['Semestre'] = df_db['semestre']
                    df_db_prep['Sexo_Num'] = df_db['sexo'].map({'Hombre': 0, 'Mujer': 1}).fillna(0)
                    df_db_prep['Sistema_Escolar_Num'] = df_db['sistema'].map({'Escolarizado': 0, 'Semiescolarizado': 1}).fillna(0)
                    df_db_prep['Promedio_General'] = df_db['promedio']
                    df_db_prep['Materias_Reprobadas'] = df_db['reprobadas']
                    df_db_prep['Calificacion_Ultima_Materia'] = df_db['calif_ultima']
                    df_db_prep['Dias_Asistencia'] = df_db['dias_asistencia']
                    df_db_prep['Asistencia_Clases'] = df_db['asistencia_clases']
                    df_db_prep['Cumplimiento_Tareas'] = df_db['cumplimiento']
                    df_db_prep['Participacion_Clase'] = df_db['participacion']
                    df_db_prep['Practicas_Programacion'] = df_db['practicas']
                    df_db_prep['Uso_Plataformas'] = df_db['uso_plataformas']
                    df_db_prep['Horas_Estudio_Semana'] = df_db['horas_estudio']
                    df_db_prep['Dias_Estudio_Semana'] = df_db['dias_estudio']
                    df_db_prep['Apoyo_Familiar'] = df_db['apoyo']
                    df_db_prep['Motivacion_Programacion'] = df_db['motivacion']
                    df_db_prep['Confianza_Aprobar'] = df_db['confianza']
                    df_db_prep['Dificultad_Materia'] = df_db['dificultad']
                    df_db_prep['Nivel_Estres'] = df_db['estres']
                    df_db_prep['Computadora_Propia'] = df_db['computadora'].map({'Sí': 1, 'No': 0}).fillna(0)
                    df_db_prep['Internet_Casa'] = df_db['internet'].map({'Sí': 1, 'No': 0}).fillna(0)
                    df_db_prep['Calidad_Internet'] = df_db['calidad_internet']
                    
                    X_custom = df_db_prep.reindex(columns=nombres_variables, fill_value=0)
                    X_custom_scaled = scaler.transform(X_custom)
                    predicciones_custom = model.predict(X_custom_scaled, verbose=0)
                    
                    resultados_custom = []
                    for i, prob in enumerate(predicciones_custom):
                        fila = df_db_prep.iloc[i]
                        prob_num = prob[0]
                        
                        if prob_num >= 0.70:
                            nivel = "🔴 Alto"
                            estado = "⚠️ RIESGO"
                        elif prob_num >= 0.40:
                            nivel = "🟡 Medio"
                            estado = "⚠️ RIESGO"
                        else:
                            nivel = "🟢 Bajo"
                            estado = "✅ ESTABLE"
                            
                        if estado == "⚠️ RIESGO":
                            impacto_variables = X_custom_scaled[i] * correlaciones
                            impacto_dict = {nombres_variables[j]: impacto_variables[j] for j in range(len(nombres_variables))}
                            top_3 = sorted(impacto_dict.items(), key=lambda x: x[1], reverse=True)[:3]
                            motivos = [f"{str(var).replace('_', ' ')} ({fila.get(var, 'N/A')})" for var, imp in top_3]
                            motivos_str = " | ".join(motivos)
                        else:
                            motivos_str = "Buen rendimiento general"
                            
                        # Guardamos info valiosa para el Dashboard
                        resultados_custom.append({
                            "Matrícula": fila.get('Matrícula'),
                            "Nombre": fila.get('Nombre_completo'),
                            "Docente Asignado": fila.get('Docente_Tutor'),
                            "Semestre": fila.get('Semestre'),
                            "Sistema_Escolar": fila.get('Sistema_Escolar_Num'),
                            "Nivel_Estres": fila.get('Nivel_Estres'),
                            "Promedio_General": fila.get('Promedio_General'),
                            "Asistencia_Clases": fila.get('Asistencia_Clases'),
                            "Horas_Estudio": fila.get('Horas_Estudio_Semana'),
                            "Resultado IA": estado,
                            "Nivel de Riesgo": nivel,
                            "Prob. Exacta (%)": f"{prob_num * 100:.2f}%",
                            "Factores Críticos": motivos_str
                        })
                        
                    df_resultados = pd.DataFrame(resultados_custom)
                    
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df_resultados.to_excel(writer, index=False, sheet_name='Diagnóstico')
                        workbook = writer.book
                        worksheet = writer.sheets['Diagnóstico']
                        
                        fill_riesgo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        fill_estable = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        
                        try:
                            idx_col_resultado = df_resultados.columns.get_loc("Resultado IA") + 1
                            for r_num in range(2, len(df_resultados) + 2):
                                cell_val = worksheet.cell(row=r_num, column=idx_col_resultado).value
                                current_fill = fill_riesgo if cell_val == "⚠️ RIESGO" else fill_estable if cell_val == "✅ ESTABLE" else None
                                if current_fill:
                                    for c_num in range(1, len(df_resultados.columns) + 1):
                                        worksheet.cell(row=r_num, column=c_num).fill = current_fill
                        except:
                            pass
                            
                    st.session_state['df_resultados'] = df_resultados
                    st.session_state['excel_data'] = buffer.getvalue()
                    st.session_state['analisis_completado'] = True
            except Exception as e:
                st.error(f"❌ ERROR DETECTADO: {str(e)}")
                traceback.print_exc()
                st.session_state['analisis_completado'] = False

        if st.session_state['analisis_completado'] and st.session_state['df_resultados'] is not None:
            st.markdown("---")
            st.subheader("📋 Diagnóstico de Alumnos Activos")
            
            cols_vista = ['Matrícula', 'Nombre', 'Semestre', 'Resultado IA', 'Nivel de Riesgo', 'Prob. Exacta (%)', 'Factores Críticos']
            df_vista_rapida = st.session_state['df_resultados'][cols_vista]
            
            df_estilado = df_vista_rapida.style.apply(colorear_filas, axis=1)
            st.dataframe(df_estilado, use_container_width=True)
            
            st.download_button(
                label="📥 Descargar Reporte de Diagnóstico (Excel)",
                data=st.session_state['excel_data'],
                file_name=f"Reporte_IA_{time.strftime('%Y%m%d-%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            

    def mostrar_dashboard_interactivo():
        st.title("Dashboard de Alumnos Registrados")
        st.markdown("Indicadores construidos con los alumnos dados de alta en el sistema, sus cuestionarios y sus evaluaciones docentes.")

        colores_proyecto = {
            "Estable": "#2ecc71",
            "Riesgo": "#ff5722",
            "No Definido": "#7f8c8d"
        }
        orden_resultado = ["Estable", "Riesgo", "No Definido"]

        @st.cache_data(ttl=15, show_spinner=False)
        def cargar_alumnos_sistema(usuario_actual, rol_actual):
            query = """
                SELECT
                    u.matricula, u.nombre,
                    (SELECT nombre FROM usuarios WHERE matricula = u.docente_id) as docente_tutor,
                    ra.semestre, ra.sexo, ra.sistema,
                    ra.horas_estudio, ra.dias_estudio, ra.apoyo, ra.motivacion,
                    ra.confianza, ra.dificultad, ra.estres, ra.computadora,
                    ra.internet, ra.calidad_internet,
                    ed.promedio, ed.reprobadas, ed.calif_ultima,
                    ed.dias_asistencia, ed.asistencia_clases, ed.cumplimiento,
                    ed.participacion, ed.practicas, ed.uso_plataformas
                FROM usuarios u
                INNER JOIN respuestas_alumnos ra ON u.matricula = ra.matricula
                INNER JOIN evaluaciones_docentes ed ON u.matricula = ed.matricula
                WHERE u.rol = 'alumno'
            """
            parametros = []
            if rol_actual == 'docente':
                query += " AND u.docente_id = %s"
                parametros.append(usuario_actual)

            query_total = "SELECT COUNT(*) AS total FROM usuarios u WHERE u.rol = 'alumno'"
            parametros_total = []
            if rol_actual == 'docente':
                query_total += " AND u.docente_id = %s"
                parametros_total.append(usuario_actual)

            with get_db_connection() as conn:
                total_registrados = pd.read_sql(query_total, conn, params=parametros_total if parametros_total else None).iloc[0]['total']
                df_db = pd.read_sql(query, conn, params=parametros if parametros else None)

            if df_db.empty:
                return pd.DataFrame(), int(total_registrados)

            df_sistema = pd.DataFrame()
            df_sistema['Matrícula'] = df_db['matricula']
            df_sistema['Nombre'] = df_db['nombre']
            df_sistema['Docente_Tutor'] = df_db['docente_tutor'].fillna('Sin asignar')
            df_sistema['Semestre'] = df_db['semestre']
            df_sistema['Sexo'] = df_db['sexo']
            df_sistema['Sistema_Escolar'] = df_db['sistema']
            df_sistema['Sistema_Escolar_Num'] = df_db['sistema'].map({'Escolarizado': 1, 'Semiescolarizado': 0}).fillna(0)
            df_sistema['Promedio_General'] = df_db['promedio']
            df_sistema['Materias_Reprobadas'] = df_db['reprobadas']
            df_sistema['Calificacion_Ultima_Materia'] = df_db['calif_ultima']
            df_sistema['Dias_Asistencia'] = df_db['dias_asistencia']
            df_sistema['Asistencia_Clases'] = df_db['asistencia_clases']
            df_sistema['Cumplimiento_Tareas'] = df_db['cumplimiento']
            df_sistema['Participacion_Clase'] = df_db['participacion']
            df_sistema['Practicas_Programacion'] = df_db['practicas']
            df_sistema['Uso_Plataformas'] = df_db['uso_plataformas']
            df_sistema['Horas_Estudio_Semana'] = df_db['horas_estudio']
            df_sistema['Dias_Estudio_Semana'] = df_db['dias_estudio']
            df_sistema['Apoyo_Familiar'] = df_db['apoyo']
            df_sistema['Motivacion_Programacion'] = df_db['motivacion']
            df_sistema['Confianza_Aprobar'] = df_db['confianza']
            df_sistema['Dificultad_Materia'] = df_db['dificultad']
            df_sistema['Nivel_Estres'] = df_db['estres']
            df_sistema['Computadora_Propia'] = df_db['computadora']
            df_sistema['Computadora_Propia_Num'] = df_db['computadora'].map({'Sí': 1, 'No': 0, 'SÃ­': 1}).fillna(0)
            df_sistema['Internet_Casa'] = df_db['internet']
            df_sistema['Internet_Casa_Num'] = df_db['internet'].map({'Sí': 1, 'No': 0, 'SÃ­': 1}).fillna(0)
            df_sistema['Calidad_Internet'] = df_db['calidad_internet']
            df_sistema['Resultado_Academico'] = np.where(
                (pd.to_numeric(df_sistema['Promedio_General'], errors='coerce') >= 70) &
                (pd.to_numeric(df_sistema['Materias_Reprobadas'], errors='coerce').fillna(0) == 0),
                'Estable',
                'Riesgo'
            )
            return df_sistema, int(total_registrados)

        def aplicar_diagnostico_ia(df_base):
            df_dash = df_base.copy()
            df_dash["Resultado"] = df_dash["Resultado_Academico"]
            fuente_dash = "Alumnos registrados en el sistema con cuestionario y evaluacion docente"

            df_ia = st.session_state.get('df_resultados')
            columnas_ia = {"Matrícula", "Resultado IA"}
            if df_ia is None or df_ia.empty or not columnas_ia.issubset(set(df_ia.columns)):
                return df_dash, fuente_dash

            columnas_merge = ["Matrícula", "Resultado IA"]
            for col in ["Nivel de Riesgo", "Prob. Exacta (%)", "Factores Críticos"]:
                if col in df_ia.columns:
                    columnas_merge.append(col)

            df_ia_merge = df_ia[columnas_merge].copy()
            df_dash["_matricula_key"] = df_dash["Matrícula"].astype(str).str.strip()
            df_ia_merge["_matricula_key"] = df_ia_merge["Matrícula"].astype(str).str.strip()
            df_dash = df_dash.merge(
                df_ia_merge.drop(columns=["Matrícula"]),
                on="_matricula_key",
                how="left"
            ).drop(columns=["_matricula_key"])

            tiene_ia = df_dash["Resultado IA"].notna()
            df_dash.loc[tiene_ia, "Resultado"] = df_dash.loc[tiene_ia, "Resultado IA"]
            if tiene_ia.any():
                fuente_dash = "Diagnostico de IA de alumnos activos con expediente completo"
            return df_dash, fuente_dash

        def normalizar_resultado(df_base):
            df_norm = df_base.copy()
            if "Resultado" in df_norm.columns:
                serie = df_norm["Resultado"]
            elif "Resultado_Exito" in df_norm.columns:
                serie = df_norm["Resultado_Exito"]
            else:
                return df_norm, False

            if pd.api.types.is_numeric_dtype(serie):
                if "Resultado_Exito" in df_norm.columns and "Resultado" not in df_norm.columns:
                    df_norm["Resultado_Cat"] = serie.map({1: "Estable", 0: "Riesgo"})
                else:
                    df_norm["Resultado_Cat"] = serie.map({0: "Estable", 1: "Riesgo"})
            else:
                limpio = serie.astype(str).str.strip().str.lower()
                df_norm["Resultado_Cat"] = limpio.map({
                    "aprobado": "Estable",
                    "aprobada": "Estable",
                    "estable": "Estable",
                    "reprobado": "Riesgo",
                    "reprobada": "Riesgo",
                    "riesgo": "Riesgo"
                })
                df_norm.loc[limpio.str.contains("estable", na=False), "Resultado_Cat"] = "Estable"
                df_norm.loc[limpio.str.contains("riesgo", na=False), "Resultado_Cat"] = "Riesgo"
            df_norm["Resultado_Cat"] = df_norm["Resultado_Cat"].fillna("No Definido")
            return df_norm, True

        def columnas_existentes(df_base, columnas):
            return [col for col in columnas if col in df_base.columns]

        def limpiar_numerica(df_base, columna):
            valores = pd.to_numeric(df_base[columna], errors="coerce")
            return df_base.assign(**{columna: valores}).dropna(subset=[columna])

        def fig_histograma(df_base, columna, titulo, etiqueta=None, bins=None):
            data = limpiar_numerica(df_base, columna)
            if data.empty:
                return None
            fig = px.histogram(
                data,
                x=columna,
                color="Resultado_Cat",
                marginal="box",
                nbins=bins,
                barmode="overlay",
                opacity=0.68,
                color_discrete_map=colores_proyecto,
                category_orders={"Resultado_Cat": orden_resultado},
                title=titulo
            )
            fig.update_layout(
                xaxis_title=etiqueta or columna.replace("_", " "),
                yaxis_title="Frecuencia (numero de estudiantes)",
                legend_title_text="Estatus del Alumno",
                bargap=0.04,
                height=430,
                margin=dict(t=60, b=35, l=35, r=20)
            )
            return fig

        def fig_barras(df_base, columna, titulo, orientacion="v", porcentaje=False):
            data = df_base.dropna(subset=[columna, "Resultado_Cat"]).copy()
            if data.empty:
                return None
            if columna == "Resultado_Cat":
                conteo = data[columna].value_counts().rename_axis("Resultado_Cat").reset_index(name="Alumnos")
                y_valor = "Alumnos"
                etiqueta_y = "Numero de estudiantes"
            elif porcentaje:
                conteo = data.groupby([columna, "Resultado_Cat"]).size().reset_index(name="Alumnos")
                total = conteo.groupby(columna)["Alumnos"].transform("sum")
                conteo["Porcentaje"] = np.where(total > 0, conteo["Alumnos"] / total * 100, 0)
                y_valor = "Porcentaje"
                etiqueta_y = "Porcentaje (%)"
            else:
                conteo = data.groupby([columna, "Resultado_Cat"]).size().reset_index(name="Alumnos")
                y_valor = "Alumnos"
                etiqueta_y = "Numero de estudiantes"

            fig_args = dict(
                data_frame=conteo,
                color="Resultado_Cat",
                color_discrete_map=colores_proyecto,
                category_orders={"Resultado_Cat": orden_resultado},
                title=titulo,
                barmode="group",
                height=430,
                text=y_valor
            )
            if orientacion == "h":
                fig = px.bar(x=y_valor, y=columna, orientation="h", **fig_args)
                fig.update_layout(xaxis_title=etiqueta_y, yaxis_title=columna.replace("_", " "))
            else:
                fig = px.bar(x=columna, y=y_valor, **fig_args)
                fig.update_layout(xaxis_title=columna.replace("_", " "), yaxis_title=etiqueta_y)
            fig.update_traces(texttemplate="%{text:.1f}" if porcentaje else "%{text}", textposition="outside")
            fig.update_layout(legend_title_text="Estatus del Alumno", margin=dict(t=60, b=35, l=35, r=20))
            return fig

        df_cargado, total_registrados = cargar_alumnos_sistema(st.session_state['usuario_actual'], st.session_state.get('rol_actual'))

        if df_cargado.empty:
            st.metric("Alumnos registrados", f"{total_registrados:,}")
            st.warning("Aún no hay alumnos con datos completos para construir las gráficas del dashboard.")
            st.info("Para aparecer aquí, el alumno debe estar registrado, responder su cuestionario y contar con evaluación docente.")
            return

        df_cargado, fuente = aplicar_diagnostico_ia(df_cargado)
        df, tiene_resultado = normalizar_resultado(df_cargado)
        if not tiene_resultado:
            st.error("No se pudo calcular el estatus de los alumnos registrados.")
            return

        total_alumnos = len(df)
        estables = int((df["Resultado_Cat"] == "Estable").sum())
        en_riesgo = int((df["Resultado_Cat"] == "Riesgo").sum())
        tasa_estables = (estables / total_alumnos) * 100 if total_alumnos else 0

        st.caption(f"Fuente: {fuente}")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Alumnos registrados", f"{total_registrados:,}")
        m2.metric("Expedientes completos", f"{total_alumnos:,}")
        m3.metric("Estables", f"{estables:,}", f"{tasa_estables:.1f}%")
        m4.metric("En riesgo", f"{en_riesgo:,}", f"{(en_riesgo / total_alumnos * 100 if total_alumnos else 0):.1f}%", delta_color="inverse")

        hist_notebook = [
            ("Promedio_General", "Distribucion del Promedio General", "Promedio General del Estudiante", 28),
            ("Asistencia_Clases", "Porcentaje de Asistencia a Clases", "Asistencia a clases", 20),
            ("Horas_Estudio_Semana", "Horas de Estudio Semanales", "Horas de estudio a la semana", 16),
            ("Materias_Reprobadas", "Historial de Materias Reprobadas", "Materias reprobadas acumuladas", 12),
            ("Calificacion_Ultima_Materia", "Calificacion de la Ultima Materia de Programacion", "Calificacion ultima materia", 20),
            ("Practicas_Programacion", "Practicas de Programacion Realizadas", "Practicas de programacion", 12),
            ("Motivacion_Programacion", "Motivacion hacia Programacion", "Motivacion", 8),
            ("Confianza_Aprobar", "Confianza Autopercibida para Aprobar", "Confianza", 8),
            ("Nivel_Estres", "Nivel de Estres Academico", "Nivel de estres", 8),
            ("Uso_Plataformas", "Uso de Plataformas Educativas", "Uso de plataformas", 8),
            ("Dificultad_Materia", "Dificultad Percibida de la Materia", "Dificultad percibida", 8),
            ("Dias_Asistencia", "Dias de Asistencia Efectiva", "Dias de asistencia", 8),
            ("Cumplimiento_Tareas", "Cumplimiento de Tareas", "Cumplimiento de tareas", 8),
            ("Participacion_Clase", "Participacion Activa en Clase", "Participacion en clase", 8),
            ("Computadora_Propia", "Disponibilidad de Computadora Propia", "Computadora propia", None),
        ]

        cat_notebook = [
            ("Sistema_Escolar", "Incidencia de Riesgo segun Modalidad", "v", True),
            ("Internet_Casa", "Conectividad a Internet y Relacion con el Exito", "v", False),
            ("Computadora_Propia", "Hardware Dedicado frente al Rezago", "v", True),
            ("Participacion_Clase", "Participacion Interactiva y Vinculo Curricular", "v", False),
            ("Cumplimiento_Tareas", "Estatus vinculado al Cumplimiento de Entregas", "v", True),
            ("Uso_Plataformas", "Interaccion con Plataformas de Aprendizaje", "h", False),
        ]

        seccion_dashboard = st.radio(
            "Vista del dashboard",
            ["Resumen", "Histogramas", "Categoricas", "Dispersion", "Correlacion"],
            horizontal=True
        )

        if seccion_dashboard == "Resumen":
            col_a, col_b = st.columns([1, 1])
            with col_a:
                fig_pie = px.pie(
                    df,
                    names="Resultado_Cat",
                    color="Resultado_Cat",
                    color_discrete_map=colores_proyecto,
                    category_orders={"Resultado_Cat": orden_resultado},
                    hole=0.45,
                    title="Distribucion General del Estatus de IA"
                )
                fig_pie.update_traces(textposition="inside", textinfo="percent+label")
                fig_pie.update_layout(height=430, margin=dict(t=60, b=25, l=20, r=20))
                st.plotly_chart(fig_pie, use_container_width=True)
            with col_b:
                fig_resultado = fig_barras(df, "Resultado_Cat", "Cantidad de Estudiantes por Estatus")
                if fig_resultado:
                    fig_resultado.update_layout(showlegend=False)
                    st.plotly_chart(fig_resultado, use_container_width=True)

            disponibles = columnas_existentes(df, [c[0] for c in hist_notebook] + [c[0] for c in cat_notebook])
            st.dataframe(
                pd.DataFrame({"Grafica integrada": disponibles}),
                use_container_width=True,
                hide_index=True
            )

        if seccion_dashboard == "Histogramas":
            st.subheader("Distribuciones de los alumnos registrados")
            graficas_hist = [(col, titulo, etiqueta, bins) for col, titulo, etiqueta, bins in hist_notebook if col in df.columns]
            if not graficas_hist:
                st.warning("No se encontraron columnas compatibles para los histogramas del notebook.")
            for idx in range(0, len(graficas_hist), 2):
                cols = st.columns(2)
                for contenedor, item in zip(cols, graficas_hist[idx:idx + 2]):
                    col, titulo, etiqueta, bins = item
                    with contenedor:
                        if pd.api.types.is_numeric_dtype(pd.to_numeric(df[col], errors="coerce")) and df[col].nunique(dropna=True) > 8:
                            fig = fig_histograma(df, col, titulo, etiqueta, bins)
                        else:
                            fig = fig_barras(df, col, titulo)
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)

        if seccion_dashboard == "Categoricas":
            st.subheader("Comparativas por características de los alumnos")
            graficas_cat = [(col, titulo, orientacion, porcentaje) for col, titulo, orientacion, porcentaje in cat_notebook if col in df.columns]
            if not graficas_cat:
                st.warning("No se encontraron columnas compatibles para las graficas categoricas.")
            for idx in range(0, len(graficas_cat), 2):
                cols = st.columns(2)
                for contenedor, item in zip(cols, graficas_cat[idx:idx + 2]):
                    col, titulo, orientacion, porcentaje = item
                    with contenedor:
                        fig = fig_barras(df, col, titulo, orientacion=orientacion, porcentaje=porcentaje)
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)

        if seccion_dashboard == "Dispersion":
            st.subheader("Relaciones entre hábitos y desempeño")
            if all(col in df.columns for col in ["Horas_Estudio_Semana", "Calificacion_Ultima_Materia"]):
                data_scatter = df.copy()
                data_scatter["Horas_Estudio_Semana"] = pd.to_numeric(data_scatter["Horas_Estudio_Semana"], errors="coerce")
                data_scatter["Calificacion_Ultima_Materia"] = pd.to_numeric(data_scatter["Calificacion_Ultima_Materia"], errors="coerce")
                data_scatter = data_scatter.dropna(subset=["Horas_Estudio_Semana", "Calificacion_Ultima_Materia"])
                fig_scatter = px.scatter(
                    data_scatter,
                    x="Horas_Estudio_Semana",
                    y="Calificacion_Ultima_Materia",
                    color="Resultado_Cat",
                    color_discrete_map=colores_proyecto,
                    category_orders={"Resultado_Cat": orden_resultado},
                    opacity=0.75,
                    title="Relacion entre Horas de Estudio y Ultima Calificacion"
                )
                fig_scatter.update_layout(
                    xaxis_title="Horas de estudio a la semana",
                    yaxis_title="Calificacion de la ultima materia",
                    height=520,
                    margin=dict(t=60, b=35, l=35, r=20)
                )
                st.plotly_chart(fig_scatter, use_container_width=True)
            else:
                st.info("Faltan Horas_Estudio_Semana o Calificacion_Ultima_Materia para la dispersion principal.")

            col_1, col_2 = st.columns(2)
            with col_1:
                if "Promedio_General" in df.columns:
                    fig = fig_histograma(df, "Promedio_General", "Histograma Premium de Promedio General", "Promedio General", 30)
                    if fig:
                        st.plotly_chart(fig, use_container_width=True)
            with col_2:
                if all(col in df.columns for col in ["Promedio_General", "Horas_Estudio_Semana"]):
                    data_prom = df.copy()
                    data_prom["Promedio_General"] = pd.to_numeric(data_prom["Promedio_General"], errors="coerce")
                    data_prom["Horas_Estudio_Semana"] = pd.to_numeric(data_prom["Horas_Estudio_Semana"], errors="coerce")
                    data_prom = data_prom.dropna(subset=["Promedio_General", "Horas_Estudio_Semana"])
                    fig = px.density_contour(
                        data_prom,
                        x="Horas_Estudio_Semana",
                        y="Promedio_General",
                        color="Resultado_Cat",
                        color_discrete_map=colores_proyecto,
                        category_orders={"Resultado_Cat": orden_resultado},
                        title="Dispersion Marginal: Horas de Estudio vs Promedio"
                    )
                    fig.update_traces(contours_coloring="fill", opacity=0.45)
                    fig.update_layout(height=430, margin=dict(t=60, b=35, l=35, r=20))
                    st.plotly_chart(fig, use_container_width=True)

        if seccion_dashboard == "Correlacion":
            st.subheader("Matriz de Correlacion de Pearson")
            df_corr = df.copy()
            if "Sistema_Escolar" in df_corr.columns and "Sistema_Escolar_Num" not in df_corr.columns:
                df_corr["Sistema_Escolar_Num"] = df_corr["Sistema_Escolar"].map({"Escolarizado": 1, "Semiescolarizado": 0})
            if "Resultado_Cat" in df_corr.columns:
                df_corr["Resultado_Exito"] = df_corr["Resultado_Cat"].map({"Estable": 1, "Riesgo": 0})

            columnas_corr_preferidas = columnas_existentes(df_corr, [
                "Promedio_General", "Sistema_Escolar_Num", "Resultado_Exito",
                "Horas_Estudio_Semana", "Materias_Reprobadas", "Calificacion_Ultima_Materia",
                "Dias_Asistencia", "Asistencia_Clases", "Cumplimiento_Tareas",
                "Participacion_Clase", "Practicas_Programacion", "Uso_Plataformas",
                "Motivacion_Programacion", "Confianza_Aprobar", "Dificultad_Materia", "Nivel_Estres"
            ])
            columnas_corr = st.multiselect(
                "Variables para correlacion:",
                options=df_corr.select_dtypes(include=[np.number]).columns.tolist(),
                default=columnas_corr_preferidas[:12]
            )
            if len(columnas_corr) < 2:
                st.warning("Selecciona al menos dos variables numericas para construir la matriz.")
            else:
                corr = df_corr[columnas_corr].apply(pd.to_numeric, errors="coerce").corr()
                fig_corr = px.imshow(
                    corr,
                    text_auto=".2f",
                    color_continuous_scale="RdBu_r",
                    zmin=-1,
                    zmax=1,
                    title="Mapa de Calor de Correlaciones"
                )
                fig_corr.update_layout(height=max(520, 38 * len(columnas_corr)), margin=dict(t=60, b=40, l=40, r=30))
                st.plotly_chart(fig_corr, use_container_width=True)

            if all(col in df_corr.columns for col in ["Promedio_General", "Sistema_Escolar_Num", "Resultado_Exito"]):
                st.markdown("#### Matriz especifica: Promedio, Sistema Escolar y Resultado")
                corr_sistema = df_corr[["Promedio_General", "Sistema_Escolar_Num", "Resultado_Exito"]].apply(pd.to_numeric, errors="coerce").corr()
                fig_sistema = px.imshow(
                    corr_sistema,
                    text_auto=".2f",
                    color_continuous_scale="RdBu_r",
                    zmin=-1,
                    zmax=1,
                    title="Desempeno Academico y Sistema Escolar"
                )
                fig_sistema.update_layout(height=420, margin=dict(t=60, b=35, l=35, r=20))
                st.plotly_chart(fig_sistema, use_container_width=True)
    # --- PANTALLA: ALUMNO ---
    def pantalla_alumno():
        col1, col2 = st.columns([2.2, 0.8])
        with col1:
            with st.container(border=True):
                st.markdown(f"<h1>👨‍🎓 Hola, {st.session_state['nombre']}</h1>", unsafe_allow_html=True)
                st.write("---")
                st.subheader("📋 Cuestionario de Hábitos y Contexto Estudiantil")
                
                with st.form(f"form_alumno_{st.session_state['form_alumno_version']}"):
                    sexo = st.selectbox("Sexo", ["Hombre", "Mujer"], index=None, placeholder="Selecciona una opción...")
                    semestre = st.number_input("Semestre actual", min_value=1, max_value=12, value=None, placeholder="Ej. 1")
                    sistema = st.selectbox("Sistema Escolar", ["Escolarizado", "Semiescolarizado"], index=None, placeholder="Selecciona una opción...")
                    
                    horas_estudio = st.number_input("Horas de Estudio a la Semana", min_value=0, max_value=168, value=None, placeholder="Ej. 5")
                    dias_estudio = st.selectbox("Días de Estudio a la Semana", [0, 1, 2, 3, 4, 5, 6, 7], index=None, placeholder="Selecciona una opción...")
                    
                    motivacion = st.selectbox("Motivación (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    confianza = st.selectbox("Confianza en Aprobar (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    dificultad = st.selectbox("Dificultad (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    apoyo = st.selectbox("Apoyo Familiar (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    estres = st.selectbox("Nivel de Estrés (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    computadora = st.selectbox("¿Computadora Propia?", ["Sí", "No"], index=None, placeholder="Selecciona una opción...")
                    internet = st.radio("¿Internet en Casa?", ["Sí", "No"], index=None)
                    calidad_internet = st.selectbox("Calidad de Internet (1 a 5)", [1, 2, 3, 4, 5], index=None, placeholder="Selecciona una opción...")
                    
                    if st.form_submit_button("Guardar Respuestas", type="primary", use_container_width=True):
                        if any(v is None for v in [sexo, semestre, sistema, horas_estudio, dias_estudio, motivacion, confianza, dificultad, apoyo, estres, computadora, internet, calidad_internet]):
                            st.error("❌ Todos los campos son obligatorios. Por favor, responde el cuestionario por completo antes de guardar.")
                        else:
                            try:
                                with get_db_connection() as conn:
                                    with conn.cursor() as c:
                                        consulta = '''REPLACE INTO respuestas_alumnos
                                                    (matricula, sexo, semestre, sistema, horas_estudio, dias_estudio, 
                                                    motivacion, confianza, dificultad, apoyo, estres, computadora, internet, calidad_internet)
                                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
                                        c.execute(consulta, (st.session_state['usuario_actual'], sexo, semestre, sistema, horas_estudio, dias_estudio, 
                                                             motivacion, confianza, dificultad, apoyo, estres, computadora, internet, calidad_internet))
                                        conn.commit()
                                st.success("ðŸŽ‰ Â¡Tus respuestas han sido guardadas con Ã©xito!")
                                st.session_state['form_alumno_version'] += 1
                                st.rerun()
                            except mysql.connector.Error as err:
                                st.error(f"Error al guardar cuestionario: {err}")
        with col2:
            with st.container(border=True):
                st.markdown("### ℹ️ Información de tu Perfil")
                st.info(f"**Matrícula:**\n{st.session_state['usuario_actual']}")
                st.info(f"**Rol Asignado:**\nAlumno")
                st.write("---")
                st.write("Asegúrate de responder de manera honesta para que el algoritmo de IA pueda estimar tu estatus de riesgo con precisión.")

    # --- PANTALLA: DOCENTE / ADMINISTRATIVO ---
    def pantalla_docente():
        col1, col2 = st.columns([2.1, 0.9])
        lista_alumnos_pendientes = []
        lista_usuarios_crud = []
        dict_docentes = {"Ninguno": None}
        
        try:
            with get_db_connection() as conn:
                with conn.cursor() as c:
                    query_alumnos = """
                        SELECT u.matricula, u.nombre, 
                               (SELECT nombre FROM usuarios WHERE matricula = u.docente_id) as nombre_maestro,
                               ra.matricula as tiene_alumno, 
                               ed.matricula as tiene_docente
                        FROM usuarios u
                        LEFT JOIN respuestas_alumnos ra ON u.matricula = ra.matricula
                        LEFT JOIN evaluaciones_docentes ed ON u.matricula = ed.matricula
                        WHERE u.rol='alumno' AND (ra.matricula IS NULL OR ed.matricula IS NULL)
                    """
                    if st.session_state['rol_actual'] == 'docente':
                        query_alumnos += " AND u.docente_id=%s"
                        c.execute(query_alumnos, (st.session_state['usuario_actual'],))
                    else:
                        c.execute(query_alumnos)
                        
                    lista_alumnos_pendientes = c.fetchall()
                    
                    if 'admin' in st.session_state['rol_actual']:
                        c.execute("SELECT matricula, nombre, rol, correo, password, docente_id FROM usuarios")
                        lista_usuarios_crud = c.fetchall()
                        c.execute("SELECT matricula, nombre FROM usuarios WHERE rol='docente'")
                        for row in c.fetchall():
                            dict_docentes[f"{row[1]} ({row[0]})"] = row[0]
                    else:
                        c.execute("SELECT matricula, nombre, rol, correo, password, docente_id FROM usuarios WHERE rol='alumno' AND docente_id=%s", (st.session_state['usuario_actual'],))
                        lista_usuarios_crud = c.fetchall()
        except mysql.connector.Error as err:
            st.error(f"Error al cargar datos desde la base de datos: {err}")

        with col1:
            with st.container(border=True):
                st.markdown(f"<h1>👨‍🏫 Panel del Personal Académico</h1>", unsafe_allow_html=True)
                st.write("---")
                
                # Menú de Tabs
                opciones_tabs = ["👥 Gestión de Usuarios (CRUD)", "🚀 Ejecutar Diagnóstico", "📊 Dashboard Interactivo"]
                if st.session_state['rol_actual'] == 'docente':
                    opciones_tabs = ["📝 Carga la información del Alumno"] + opciones_tabs
                    
                cols_menu = st.columns(len(opciones_tabs))
                for idx, opcion in enumerate(opciones_tabs):
                    with cols_menu[idx]:
                        color_tipo = "primary" if st.session_state['tab_actual'] == opcion else "secondary"
                        if st.button(opcion, type=color_tipo, use_container_width=True, key=f"nav_{opcion}"):
                            st.session_state['tab_actual'] = opcion
                            st.rerun()
                            
                st.write("---")
                
                # --- VISTA: CARGA LA INFORMACIÓN DEL ALUMNO ---
                if st.session_state['tab_actual'] == "📝 Carga la información del Alumno" and st.session_state['rol_actual'] == 'docente':
                    st.subheader("Registro de Desempeño Académico")
                    with st.form(f"form_docente_{st.session_state['form_docente_version']}"):
                        matricula_ingresada = st.text_input("Matrícula del Alumno a Evaluar", value=st.session_state['alumno_seleccionado_evaluar']).strip()
                        
                        c_1, c_2, c_3 = st.columns(3)
                        with c_1: promedio = st.number_input("Promedio General", min_value=0.0, max_value=100.0, value=0.0)
                        with c_2: reprobadas = st.number_input("Materias Reprobadas", min_value=0, value=0)
                        with c_3: calif_ultima = st.number_input("Calificación Última Materia", min_value=0, max_value=100, value=0)
                        
                        asistencia_clases = st.slider("Asistencia (1-5)", 1, 5, 1)
                        cumplimiento = st.slider("Cumplimiento (1-5)", 1, 5, 1)
                        participacion = st.slider("Participación (1-5)", 1, 5, 1)
                        practicas = st.slider("Prácticas (1-5)", 1, 5, 1)
                        uso_plataformas = st.slider("Uso Plataformas (1-5)", 1, 5, 1)
                        
                        dias_asistencia = st.number_input("Días Totales Asistidos a la Semana", min_value=0, max_value=7, value=0)
                        
                        if st.form_submit_button("Actualizar Expediente Escolar", type="primary", use_container_width=True):
                            if not matricula_ingresada:
                                st.error("❌ Debes escribir una matrícula.")
                            else:
                                try:
                                    with get_db_connection() as conn:
                                        with conn.cursor() as c:
                                            c.execute("SELECT nombre, rol, docente_id FROM usuarios WHERE matricula = %s", (matricula_ingresada,))
                                            usuario_encontrado = c.fetchone()
                                            
                                            if not usuario_encontrado:
                                                st.error("❌ La matrícula no existe.")
                                            elif usuario_encontrado[2] != st.session_state['usuario_actual']:
                                                st.error("❌ Este alumno no está asignado bajo tu cargo.")
                                            else:
                                                consulta = '''REPLACE INTO evaluaciones_docentes
                                                            (matricula, promedio, reprobadas, calif_ultima, dias_asistencia, 
                                                            asistencia_clases, cumplimiento, participacion, practicas, uso_plataformas)
                                                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
                                                c.execute(consulta, (matricula_ingresada, promedio, reprobadas, calif_ultima, dias_asistencia, 
                                                                     asistencia_clases, cumplimiento, participacion, practicas, uso_plataformas))
                                                conn.commit()
                                                st.success(f"ðŸŽ‰ Â¡Expediente de {usuario_encontrado[0]} guardado!")
                                                st.session_state['alumno_seleccionado_evaluar'] = ""
                                                st.session_state['form_docente_version'] += 1
                                                st.rerun()
                                except mysql.connector.Error as err:
                                    st.error(f"Error al guardar evaluación: {err}")

                # --- VISTA: GESTIÓN DE USUARIOS (CRUD) ---
                elif st.session_state['tab_actual'] == "👥 Gestión de Usuarios (CRUD)":
                    st.subheader("👥 Control y Gestión Institucional de Usuarios")
                    dict_usuarios_completo = {f"{row[1]} ({row[0]}) - [{row[2].upper()}]": row for row in lista_usuarios_crud}
                    
                    col_c1, col_c2, col_c3 = st.columns(3)
                    
                    with col_c1:
                        with st.expander("➕ Registrar Nuevo Usuario", expanded=False):
                            if 'admin' in st.session_state['rol_actual']:
                                al_rol = st.selectbox("Asignar Rol", ["alumno", "docente", "administrative"], key="alta_rol_usuario", format_func=lambda rol: {"alumno": "Alumno", "docente": "Docente", "administrative": "Administrador"}.get(rol, rol))
                            else:
                                al_rol = "alumno"

                            with st.form(f"form_alta_global_{st.session_state['form_alta_usuario_version']}"):
                                label_u = "Matrícula / Usuario" if 'admin' in st.session_state['rol_actual'] else "Matrícula del Alumno"
                                al_matricula = st.text_input(label_u).strip()
                                al_nombre = st.text_input("Nombre Completo")
                                al_correo = st.text_input("Correo Electrónico")
                                al_password = st.text_input("Contraseña por Defecto", value="", placeholder="Ej. Temporal123*")
                                
                                if al_rol == "alumno":
                                    doc_asig = st.selectbox("Docente Tutor", list(dict_docentes.keys()))
                                    al_docente_id = dict_docentes[doc_asig]
                                else:
                                    al_docente_id = None
                                    
                                if st.form_submit_button("Guardar Usuario", type="primary", use_container_width=True):
                                    if not al_matricula or not al_nombre:
                                        st.error("❌ Matrícula y Nombre Obligatorios.")
                                    elif not validar_password_moodle(al_password):
                                        st.error("❌ La contraseña debe tener al menos 8 caracteres, 1 mayúscula, 1 minúscula, 1 número y 1 carácter especial.")
                                    else:
                                        try:
                                            password_hash = generar_md5(al_password)
                                            with get_db_connection() as conn:
                                                with conn.cursor() as c:
                                                    c.execute("INSERT INTO usuarios (matricula, password, rol, nombre, correo, docente_id) VALUES (%s, %s, %s, %s, %s, %s)",
                                                              (al_matricula, password_hash, al_rol, al_nombre, al_correo, al_docente_id))
                                                    conn.commit()
                                            st.success("ðŸŽ‰ Usuario dado de alta exitosamente.")
                                            st.session_state['form_alta_usuario_version'] += 1
                                            st.rerun()
                                        except mysql.connector.Error as err:
                                            st.error(f"Error: {err}")

                    with col_c2:
                        with st.expander("📝 Editar Usuario Seleccionado", expanded=False):
                            if not dict_usuarios_completo:
                                st.write("No hay usuarios disponibles.")
                            else:
                                seleccionado_edit = st.selectbox("Buscar usuario a modificar:", list(dict_usuarios_completo.keys()), key="sel_crud_edit")
                                datos_originales = dict_usuarios_completo[seleccionado_edit]
                                limpiar_edicion = st.session_state.pop('limpiar_edicion_usuario', False)
                                
                                with st.form(f"form_edicion_global_{st.session_state['form_edicion_usuario_version']}"):
                                    edit_nombre = st.text_input("Modificar Nombre Completo", value="" if limpiar_edicion else datos_originales[1])
                                    edit_correo = st.text_input("Modificar Correo", value="" if limpiar_edicion else datos_originales[3])
                                    edit_password = st.text_input("Modificar Contraseña (o dejar el Hash)", value="" if limpiar_edicion else datos_originales[4])
                                    
                                    if 'admin' in st.session_state['rol_actual']:
                                        roles_disp = ["alumno", "docente", "administrative"]
                                        if str(datos_originales[0]).lower() == "admin":
                                            roles_disp = ["administrative"]
                                        idx_r = roles_disp.index(datos_originales[2]) if datos_originales[2] in roles_disp else 0
                                        edit_rol = st.selectbox("Modificar Rol", roles_disp, index=idx_r, format_func=lambda rol: {"alumno": "Alumno", "docente": "Docente", "administrative": "Administrador"}.get(rol, rol))
                                        
                                        if edit_rol == "alumno":
                                            idx_d = 0
                                            keys_doc = list(dict_docentes.keys())
                                            for pos, k in enumerate(keys_doc):
                                                if dict_docentes[k] == datos_originales[5]:
                                                    idx_d = pos
                                                    break
                                            edit_doc_asig = st.selectbox("Modificar Docente Tutor", keys_doc, index=idx_d)
                                            edit_docente_id = dict_docentes[edit_doc_asig]
                                        else:
                                            edit_docente_id = None
                                    else:
                                        edit_rol = "alumno"
                                        edit_docente_id = st.session_state['usuario_actual']
                                        
                                    if st.form_submit_button("Actualizar Cambios", type="primary", use_container_width=True):
                                        pwd_a_guardar = edit_password
                                        valido = True
                                        
                                        if edit_password != datos_originales[4]:
                                            if not validar_password_moodle(edit_password):
                                                st.error("❌ La nueva contraseña debe tener 8 caracteres, mayúscula, minúscula, número y especial.")
                                                valido = False
                                            else:
                                                pwd_a_guardar = generar_md5(edit_password)
                                                
                                        if valido:
                                            try:
                                                with get_db_connection() as conn:
                                                    with conn.cursor() as c:
                                                        c.execute("""UPDATE usuarios 
                                                                    SET nombre=%s, correo=%s, password=%s, rol=%s, docente_id=%s 
                                                                    WHERE matricula=%s""",
                                                                  (edit_nombre, edit_correo, pwd_a_guardar, edit_rol, edit_docente_id, datos_originales[0]))
                                                        conn.commit()
                                                st.success("ðŸŽ‰ Datos de usuario actualizados.")
                                                st.session_state['limpiar_edicion_usuario'] = True
                                                st.session_state['form_edicion_usuario_version'] += 1
                                                st.rerun()
                                            except mysql.connector.Error as err:
                                                st.error(f"Error al actualizar: {err}")

                    with col_c3:
                        with st.expander("🗑️ Eliminar Usuario", expanded=False):
                            if not dict_usuarios_completo:
                                st.write("No hay registros.")
                            else:
                                seleccionado_del = st.selectbox("Buscar usuario a remover:", list(dict_usuarios_completo.keys()), key="sel_crud_del")
                                datos_eliminar = dict_usuarios_completo[seleccionado_del]
                                
                                st.warning(f"¿Remover a {datos_eliminar[1]} ({datos_eliminar[0]})? Se eliminarán en cascada sus encuestas y calificaciones.")
                                with st.form("form_baja_global"):
                                    if st.form_submit_button("❌ Confirmar Eliminación Absoluta", type="primary", use_container_width=True):
                                        if datos_eliminar[0] == st.session_state['usuario_actual']:
                                            st.error("No es posible auto-eliminarse del sistema.")
                                        else:
                                            try:
                                                with get_db_connection() as conn:
                                                    with conn.cursor() as c:
                                                        c.execute("DELETE FROM usuarios WHERE matricula=%s", (datos_eliminar[0],))
                                                        conn.commit()
                                                st.success("🗑️ Registro revocado con éxito.")
                                                st.rerun()
                                            except mysql.connector.Error as err:
                                                st.error(f"Error al eliminar: {err}")
                                                
                    st.write("---")
                    
                    if lista_usuarios_crud:
                        df_crud_vista = pd.DataFrame(lista_usuarios_crud, columns=["Matrícula", "Nombre", "Rol", "Correo", "Contraseña", "ID Docente Asignado"])
                        
                        buffer_crud = io.BytesIO()
                        with pd.ExcelWriter(buffer_crud, engine='openpyxl') as writer:
                            df_crud_vista.to_excel(writer, index=False, sheet_name='Usuarios Registrados')
                            
                        col_tit_tabla, col_btn_tabla = st.columns([8.5, 1.5])
                        with col_tit_tabla:
                            st.write("### 📋 Vista General de la Tabla de Usuarios")
                        with col_btn_tabla:
                            st.download_button(
                                label="📥 Excel",
                                data=buffer_crud.getvalue(),
                                file_name=f"Vista_General_Usuarios_{time.strftime('%Y%m%d-%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            
                        st.dataframe(df_crud_vista, use_container_width=True)
                    else:
                        st.write("### 📋 Vista General de la Tabla de Usuarios")
                        st.info("No existen usuarios registrados bajo este criterio.")

                # --- VISTA: EJECUTAR DIAGNÓSTICO ---
                elif st.session_state['tab_actual'] == "🚀 Ejecutar Diagnóstico":
                    mostrar_modulo_ia()

                # --- VISTA: DASHBOARD INTERACTIVO ---
                elif st.session_state['tab_actual'] == "📊 Dashboard Interactivo":
                    mostrar_dashboard_interactivo()

        with col2:
            with st.container(border=True):
                st.markdown("### 📋 Alumnos Pendientes")
                if lista_alumnos_pendientes:
                    for row in lista_alumnos_pendientes:
                        tiene_alumno = row[3] is not None
                        tiene_docente = row[4] is not None
                        
                        if not tiene_alumno and not tiene_docente:
                            msg_pendiente = "⏳ Pendiente: Alumno y Docente"
                            color_tag = "#ffb3b3"
                        elif not tiene_alumno:
                            msg_pendiente = "📝 Pendiente: Cuestionario Alumno"
                            color_tag = "#ffe6cc"
                        else:
                            msg_pendiente = "📊 Pendiente: Evaluación Docente"
                            color_tag = "#e6f2ff"
                            
                        if 'admin' in st.session_state['rol_actual']:
                            nombre_tutor = row[2] if row[2] else "Sin asignar"
                            st.markdown(f"""
                            <div style='padding:10px; border:1px solid #ddd; border-radius:5px; margin-bottom:8px; background-color:#fff;'>
                                <b>👤 Alumno:</b> {row[1]} (<small>{row[0]}</small>)<br>
                                <b>👨‍🏫 Docente:</b> {nombre_tutor}<br>
                                <span style='background-color:{color_tag}; padding:2px 6px; border-radius:4px; font-size:12px; font-weight:bold; display:inline-block; margin-top:4px;'>{msg_pendiente}</span>
                            </div>
                            """, unsafe_allow_html=True)
                        else:
                            if not tiene_docente:
                                if st.button(f"👤 {row[1]} ({row[0]})", key=f"btn_{row[0]}", use_container_width=True):
                                    st.session_state['alumno_seleccionado_evaluar'] = row[0]
                                    st.session_state['tab_actual'] = "📝 Carga la información del Alumno"
                                    st.rerun()
                            else:
                                st.markdown(f"<div style='padding:5px; text-align:center; font-weight:bold;'>👤 {row[1]} ({row[0]})</div>", unsafe_allow_html=True)
                            
                            st.markdown(f"<p style='text-align:center; background-color:{color_tag}; font-weight:bold; font-size:12px; margin-top:-6px; border-radius:4px;'>{msg_pendiente}</p>", unsafe_allow_html=True)
                else:
                    st.success("🎉 ¡No quedan alumnos pendientes en este periodo!")

    # --- RUTEO AUTOMÁTICO DE INTERFAZ SEGÚN EL ROL DE SESIÓN ---
    if st.session_state['rol_actual'] == 'alumno':
        pantalla_alumno()
    elif st.session_state['rol_actual'] == 'docente' or 'admin' in str(st.session_state['rol_actual']).lower():
        pantalla_docente()
